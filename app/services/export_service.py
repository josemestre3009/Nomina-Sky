"""
Servicio de exportación.
Genera archivos PDF y Excel con resúmenes de nómina detallados (incluyendo valor por día).
"""
import io
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def generar_pdf(resumenes, fecha_inicio, fecha_fin):
    """
    Genera un PDF con el resumen de nómina detallado.

    Args:
        resumenes: Lista de resúmenes generados por nomina_service
        fecha_inicio: Fecha inicio del rango
        fecha_fin: Fecha fin del rango

    Returns:
        BytesIO con el PDF generado
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )

    # Estilos
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        'TituloNomina',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1a237e')
    )
    subtitulo_style = ParagraphStyle(
        'SubtituloNomina',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#455a64')
    )
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4
    )
    total_style = ParagraphStyle(
        'Total',
        parent=styles['Normal'],
        fontSize=12,
        spaceBefore=8,
        textColor=colors.HexColor('#1b5e20'),
        alignment=TA_RIGHT
    )

    elements = []

    # Encabezado
    elements.append(Paragraph('RESUMEN DE NÓMINA', titulo_style))
    elements.append(Paragraph(
        f'Período: {fecha_inicio.strftime("%d/%m/%Y")} — {fecha_fin.strftime("%d/%m/%Y")}',
        subtitulo_style
    ))
    elements.append(Spacer(1, 20))

    for resumen in resumenes:
        emp = resumen['empleado']

        # Datos del empleado
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1a237e')))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f'<b>Empleado:</b> {emp.nombre}', info_style))
        elements.append(Paragraph(f'<b>Cédula:</b> {emp.cedula}', info_style))
        elements.append(Paragraph(f'<b>Cargo:</b> {emp.cargo}', info_style))
        elements.append(Paragraph(
            f'<b>Días Trabajados:</b> {resumen["dias_trabajados"]}',
            info_style
        ))
        elements.append(Spacer(1, 10))

        # Tabla de actividades
        if resumen['reportes']:
            data = [['Fecha', 'Actividad Realizada', 'Valor Día']]
            for r in resumen['reportes']:
                texto_act = r.actividad or ''
                if r.estado_pago == 'ausente':
                    texto_act = f"No laboró - {texto_act}"
                data.append([
                    r.fecha.strftime('%d/%m/%Y'),
                    Paragraph(texto_act[:200], styles['Normal']),
                    f'${r.valor_dia_aplicado:,.0f} COP'
                ])

            col_widths = [80, 320, 100]
            tabla = Table(data, colWidths=col_widths, repeatRows=1)
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdbdbd')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))
            elements.append(tabla)

        elements.append(Spacer(1, 10))

        # Bonos
        if resumen['bonos']:
            elements.append(Paragraph('<b>Bonos:</b>', info_style))
            for bono in resumen['bonos']:
                elements.append(Paragraph(
                    f'  • {bono.descripcion}: ${bono.valor:,.0f} COP',
                    info_style
                ))
            elements.append(Spacer(1, 6))

        # Total final (SIN desglose por día)
        elements.append(Paragraph(
            f'<b>TOTAL A PAGAR: ${resumen["total_final"]:,.0f} COP</b>',
            total_style
        ))
        elements.append(Spacer(1, 20))

    # Pie de página con fecha de generación
    elements.append(Spacer(1, 30))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    fecha_gen = ParagraphStyle('FechaGen', parent=styles['Normal'], fontSize=8,
                                textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Paragraph(
        f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} — Actividades Sky',
        fecha_gen
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_excel(resumenes, fecha_inicio, fecha_fin):
    """
    Genera un archivo Excel con el resumen de nómina detallado.

    Returns:
        BytesIO con el archivo Excel
    """
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen Nómina'

    # Estilos
    titulo_font = Font(name='Calibri', size=16, bold=True, color='1A237E')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    info_font = Font(name='Calibri', size=11)
    total_font = Font(name='Calibri', size=12, bold=True, color='1B5E20')
    bono_font = Font(name='Calibri', size=11, italic=True, color='E65100')
    border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )
    center = Alignment(horizontal='center', vertical='center')
    wrap = Alignment(wrap_text=True, vertical='top')

    # Título
    ws.merge_cells('A1:C1')
    ws['A1'] = 'RESUMEN DE NÓMINA'
    ws['A1'].font = titulo_font
    ws['A1'].alignment = center

    ws.merge_cells('A2:C2')
    ws['A2'] = f'Período: {fecha_inicio.strftime("%d/%m/%Y")} — {fecha_fin.strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(name='Calibri', size=11, color='455A64')
    ws['A2'].alignment = center

    row = 4

    for resumen in resumenes:
        emp = resumen['empleado']

        # Info del empleado
        ws.cell(row=row, column=1, value='Empleado:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=emp.nombre).font = info_font
        row += 1
        ws.cell(row=row, column=1, value='Cédula:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=emp.cedula).font = info_font
        row += 1
        ws.cell(row=row, column=1, value='Cargo:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=emp.cargo).font = info_font
        row += 1
        ws.cell(row=row, column=1, value='Días Trabajados:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=resumen['dias_trabajados']).font = info_font
        row += 2

        # Headers de tabla
        headers = ['Fecha', 'Actividad Realizada', 'Valor Día']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center
        row += 1

        # Datos
        for r in resumen['reportes']:
            ws.cell(row=row, column=1, value=r.fecha.strftime('%d/%m/%Y')).border = border
            ws.cell(row=row, column=1).alignment = center
            
            texto_act = r.actividad or ''
            if r.estado_pago == 'ausente':
                texto_act = f"No laboró - {texto_act}"

            cell_act = ws.cell(row=row, column=2, value=texto_act)
            cell_act.border = border
            cell_act.alignment = wrap
            
            cell_val = ws.cell(row=row, column=3, value=f'${r.valor_dia_aplicado:,.0f} COP')
            cell_val.border = border
            cell_val.alignment = Alignment(horizontal='right', vertical='top')
            row += 1

        row += 1

        # Bonos
        if resumen['bonos']:
            for bono in resumen['bonos']:
                ws.cell(row=row, column=1, value=f'Bono: {bono.descripcion}').font = bono_font
                ws.cell(row=row, column=2, value=f'${bono.valor:,.0f} COP').font = bono_font
                row += 1
            row += 1

        # Total final
        ws.cell(row=row, column=1, value='TOTAL A PAGAR:').font = total_font
        ws.cell(row=row, column=2, value=f'${resumen["total_final"]:,.0f} COP').font = total_font
        row += 3

    # Ajustar anchos
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20

    # Pie
    ws.cell(row=row, column=1,
            value=f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}').font = Font(
        size=8, color='999999')

    wb.save(buffer)
    buffer.seek(0)
    return buffer
